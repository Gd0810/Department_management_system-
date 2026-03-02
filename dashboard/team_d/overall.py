from decimal import Decimal
from io import BytesIO
import os

from django.http import HttpResponse, JsonResponse
from django.db.models import Sum
from django.utils import timezone

from ..models import Project, ProjectMember


def _calculate_project_payments(project):
    if not project.amount:
        return {}

    members = list(project.members.select_related("worker"))
    gold = [m for m in members if m.contribution == "gold"]
    silver = [m for m in members if m.contribution == "silver"]
    copper = [m for m in members if m.contribution == "copper"]
    total_amount = Decimal(project.amount)
    payments = {}

    if gold and not silver and not copper:
        share = total_amount / len(gold)
        for member in gold:
            payments[member.id] = share
    elif gold and silver and not copper:
        gold_total = total_amount * Decimal("0.60")
        silver_total = total_amount * Decimal("0.40")
        for member in gold:
            payments[member.id] = gold_total / len(gold)
        for member in silver:
            payments[member.id] = silver_total / len(silver)
    elif gold and copper and not silver:
        gold_total = total_amount * Decimal("0.70")
        copper_total = total_amount * Decimal("0.30")
        for member in gold:
            payments[member.id] = gold_total / len(gold)
        for member in copper:
            payments[member.id] = copper_total / len(copper)
    else:
        weight_map = {"gold": 3, "silver": 2, "copper": 1}
        total_weight = sum(weight_map[member.contribution] for member in members)
        for member in members:
            share = (Decimal(weight_map[member.contribution]) / Decimal(total_weight)) * total_amount
            payments[member.id] = share

    return payments


def build_team_report_data(dept):
    workers = dept.workers.all().order_by("name")
    projects = dept.projects.all().prefetch_related("members__worker")

    staff_count = workers.filter(worker_type="staff").count()
    intern_count = workers.filter(worker_type="intern").count()
    total_workers = workers.count()
    total_income = (
        projects.aggregate(total=Sum("amount"))["total"]
        or Decimal("0.00")
    )
    revenue_per_worker = (
        (Decimal(total_income) / Decimal(total_workers))
        if total_workers > 0
        else Decimal("0.00")
    )

    worker_income_map = {worker.id: Decimal("0.00") for worker in workers}
    for project in projects:
        payments = _calculate_project_payments(project)
        for member in project.members.all():
            worker_income_map[member.worker_id] = worker_income_map.get(member.worker_id, Decimal("0.00")) + payments.get(member.id, Decimal("0.00"))

    rows = []
    for worker in workers:
        rows.append(
            {
                "name": worker.name,
                "email": worker.email or "-",
                "date_of_join": worker.date_of_join.strftime("%Y-%m-%d"),
                "posting": worker.posting,
                "income_by_user": worker_income_map.get(worker.id, Decimal("0.00")),
            }
        )

    return {
        "department_name": dept.name,
        "total_staff_count": staff_count,
        "total_intern_count": intern_count,
        "revenue_per_worker": revenue_per_worker,
        "rows": rows,
        "generated_at": timezone.localtime().strftime("%Y-%m-%d %H:%M:%S"),
    }


def generate_team_csv_report(dept):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        has_openpyxl = True
    except ImportError:
        has_openpyxl = False

    if not has_openpyxl:
        return JsonResponse(
            {"detail": "Excel generation dependency missing. Install openpyxl."},
            status=500,
        )

    report = build_team_report_data(dept)

    wb = Workbook()
    ws = wb.active
    ws.title = "Team Report"

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 16

    summary_rows = [
        ("Department Name", report["department_name"]),
        ("Total Staff Count", str(report["total_staff_count"])),
        ("Total Intern Count", str(report["total_intern_count"])),
        ("Revenue per Worker", f"Rs {report['revenue_per_worker']:,.2f}"),
        ("Generated At", report["generated_at"]),
    ]

    row_idx = 1
    for label, value in summary_rows:
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)
        row_idx += 1

    row_idx += 1
    headers = ["Name", "Email", "Date Of Join", "Posting", "Income By User"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    row_idx += 1

    for item in report["rows"]:
        ws.cell(row=row_idx, column=1, value=item["name"])
        ws.cell(row=row_idx, column=2, value=item["email"])
        ws.cell(row=row_idx, column=3, value=item["date_of_join"])
        ws.cell(row=row_idx, column=4, value=item["posting"])
        ws.cell(row=row_idx, column=5, value=f"Rs {Decimal(item['income_by_user']):,.2f}")
        for col_idx in range(1, 6):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="left", vertical="center", wrap_text=(col_idx == 2))
        row_idx += 1

    ws.freeze_panes = "A7"

    filename = "team_overall_report.xlsx"
    output = BytesIO()
    wb.save(output)
    payload = output.getvalue()
    output.close()

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write(payload)
    return response


def generate_team_pdf_report(dept):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    except ImportError:
        return HttpResponse(
            "PDF generation dependency is missing. Install reportlab.",
            status=500,
            content_type="text/plain",
        )

    report = build_team_report_data(dept)

    font_name = "Helvetica"
    bold_font_name = "Helvetica-Bold"
    for path in [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/calibri.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("ReportSans", path))
                font_name = "ReportSans"
                bold_font_name = "ReportSans"
                break
            except Exception:
                pass

    BRAND_DARK = colors.HexColor("#1A2B4A")
    BRAND_MID = colors.HexColor("#2E5FA3")
    BRAND_LIGHT = colors.HexColor("#E8F0FB")
    ACCENT = colors.HexColor("#F0A500")
    WHITE = colors.white
    GREY_TEXT = colors.HexColor("#4A4A4A")
    ROW_ALT = colors.HexColor("#EEF3FB")

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Team Overall Report",
        author=report["department_name"],
    )
    page_w = A4[0] - 20 * mm

    def style(name, **kwargs):
        return ParagraphStyle(name, **kwargs)

    getSampleStyleSheet()

    S_REPORT_TITLE = style(
        "ReportTitle",
        fontName=bold_font_name, fontSize=22,
        textColor=WHITE, alignment=TA_LEFT,
        spaceAfter=0, spaceBefore=0, leading=24,
    )
    S_REPORT_SUBTITLE = style(
        "ReportSubtitle",
        fontName=bold_font_name, fontSize=16,
        textColor=ACCENT, alignment=TA_RIGHT,
        spaceAfter=0, spaceBefore=0, leading=18,
    )
    S_SECTION_HEADING = style(
        "SectionHeading",
        fontName=bold_font_name, fontSize=11,
        textColor=BRAND_DARK, spaceBefore=14, spaceAfter=4,
    )
    S_LABEL = style(
        "Label",
        fontName=bold_font_name, fontSize=9,
        textColor=BRAND_DARK,
    )
    S_TABLE_HEADER = style(
        "TableHeader",
        fontName=bold_font_name, fontSize=9,
        textColor=WHITE, alignment=TA_LEFT,
    )
    S_TABLE_CELL = style(
        "TableCell",
        fontName=bold_font_name, fontSize=9,
        textColor=GREY_TEXT, leading=12,
    )
    S_TABLE_CELL_BOLD = style(
        "TableCellBold",
        fontName=bold_font_name, fontSize=8.5,
        textColor=BRAND_DARK, leading=12,
    )

    header_table = Table(
        [[
            Paragraph("Team Overall Report", S_REPORT_TITLE),
            Paragraph(report["department_name"], S_REPORT_SUBTITLE),
        ]],
        colWidths=[page_w * 0.7, page_w * 0.3],
    )
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), BRAND_DARK),
        ("TOPPADDING", (0, 0), (-1, -1), 14),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
        ("LEFTPADDING", (0, 0), (-1, -1), 16),
        ("RIGHTPADDING", (0, 0), (-1, -1), 16),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, 0), "LEFT"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))

    accent_bar = Table([[""]], colWidths=[page_w], rowHeights=[4])
    accent_bar.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), ACCENT),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    def summary_card(label, value):
        return [
            Paragraph(label, S_LABEL),
            Paragraph(
                str(value),
                style(
                    "CardVal",
                    fontName=bold_font_name,
                    fontSize=16,
                    textColor=BRAND_MID,
                    spaceAfter=0,
                ),
            ),
        ]

    summary_table = Table(
        [[
            summary_card("Total Staff Count", str(report["total_staff_count"])),
            summary_card("Total Intern Count", str(report["total_intern_count"])),
            summary_card("Revenue per Worker", f"₹{report['revenue_per_worker']:,.2f}"),
        ]],
        colWidths=[page_w / 3] * 3,
    )
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), BRAND_LIGHT),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ("LEFTPADDING", (0, 0), (-1, -1), 14),
        ("RIGHTPADDING", (0, 0), (-1, -1), 14),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LINEAFTER", (0, 0), (1, 0), 0.5, colors.HexColor("#B0C4E8")),
    ]))

    rows = [[
        Paragraph("#", S_TABLE_HEADER),
        Paragraph("Name", S_TABLE_HEADER),
        Paragraph("Email", S_TABLE_HEADER),
        Paragraph("Date Of Join", S_TABLE_HEADER),
        Paragraph("Posting", S_TABLE_HEADER),
        Paragraph("Income By User", S_TABLE_HEADER),
    ]]
    for item in report["rows"]:
        rows.append([
            Paragraph(str(len(rows)), S_TABLE_CELL),
            Paragraph(item["name"], S_TABLE_CELL_BOLD),
            Paragraph(item["email"], S_TABLE_CELL),
            Paragraph(item["date_of_join"], S_TABLE_CELL),
            Paragraph(item["posting"], S_TABLE_CELL),
            Paragraph(f"₹{Decimal(item['income_by_user']):,.2f}", S_TABLE_CELL),
        ])

    row_bg_cmds = []
    for r in range(1, len(rows)):
        row_bg_cmds.append(("BACKGROUND", (0, r), (-1, r), ROW_ALT if r % 2 == 0 else WHITE))

    table = Table(
        rows,
        colWidths=[
            page_w * 0.05,
            page_w * 0.20,
            page_w * 0.28,
            page_w * 0.13,
            page_w * 0.16,
            page_w * 0.18,
        ],
        repeatRows=1,
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_MID),
        ("TOPPADDING", (0, 0), (-1, 0), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 1), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#D0DCF0")),
        ("LINEBELOW", (0, 0), (-1, 0), 0, WHITE),
        ("BOX", (0, 0), (-1, -1), 0.8, BRAND_MID),
        *row_bg_cmds,
    ]))

    footer_note = Paragraph(
        f"This report was automatically generated for the "
        f"<b>{report['department_name']}</b> department. "
        f"All figures are indicative and subject to change.",
        style(
            "Footer",
            fontName=font_name,
            fontSize=8,
            textColor=colors.HexColor("#888888"),
            alignment=TA_CENTER,
        ),
    )

    def add_page_footer(canvas_obj, doc_obj):
        canvas_obj.saveState()
        canvas_obj.setFont(font_name, 8)
        canvas_obj.setFillColor(colors.HexColor("#AAAAAA"))
        canvas_obj.drawString(
            10 * mm, 10 * mm,
            f"Team Overall Report  |  {report['department_name']}"
        )
        canvas_obj.drawRightString(
            A4[0] - 10 * mm, 10 * mm,
            f"Page {doc_obj.page}"
        )
        canvas_obj.setStrokeColor(colors.HexColor("#DDDDDD"))
        canvas_obj.line(10 * mm, 13 * mm, A4[0] - 10 * mm, 13 * mm)
        canvas_obj.restoreState()

    story = [
        header_table,
        accent_bar,
        Spacer(1, 10),
        Paragraph("Summary Overview", S_SECTION_HEADING),
        summary_table,
        Spacer(1, 14),
        HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#D0DCF0")),
        Spacer(1, 6),
        Paragraph("Team Details", S_SECTION_HEADING),
        table,
        Spacer(1, 20),
        HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#DDDDDD")),
        Spacer(1, 6),
        footer_note,
    ]
    doc.build(story, onFirstPage=add_page_footer, onLaterPages=add_page_footer)

    payload = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="team_overall_report.pdf"'
    response.write(payload)
    return response
