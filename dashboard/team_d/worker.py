from decimal import Decimal
from io import BytesIO
import os

from django.http import HttpResponse, JsonResponse
from django.utils import timezone

def _calculate_project_payments(project):
    if not project.amount:
        return {}

    members = list(project.members.select_related("worker"))
    gold = [m for m in members if m.contribution == "gold"]
    silver = [m for m in members if m.contribution == "silver"]
    copper = [m for m in members if m.contribution == "copper"]
    total_amount = Decimal(project.amount)
    payments = {}

    if gold and not silver and not copper:
        share = total_amount / len(gold)
        for member in gold:
            payments[member.id] = share
    elif gold and silver and not copper:
        gold_total = total_amount * Decimal("0.60")
        silver_total = total_amount * Decimal("0.40")
        for member in gold:
            payments[member.id] = gold_total / len(gold)
        for member in silver:
            payments[member.id] = silver_total / len(silver)
    elif gold and copper and not silver:
        gold_total = total_amount * Decimal("0.70")
        copper_total = total_amount * Decimal("0.30")
        for member in gold:
            payments[member.id] = gold_total / len(gold)
        for member in copper:
            payments[member.id] = copper_total / len(copper)
    else:
        weight_map = {"gold": 3, "silver": 2, "copper": 1}
        total_weight = sum(weight_map[member.contribution] for member in members)
        for member in members:
            share = (Decimal(weight_map[member.contribution]) / Decimal(total_weight)) * total_amount
            payments[member.id] = share

    return payments


def build_worker_report_data(dept, worker):
    assigned_projects = list(
        dept.projects.filter(members__worker=worker)
        .distinct()
        .prefetch_related("members__worker")
        .order_by("-start_date", "-id")
    )

    total_income = Decimal("0.00")
    finished_count = 0
    project_amount_sum = Decimal("0.00")
    amount_count = 0

    for project in assigned_projects:
        payments = _calculate_project_payments(project)
        for member in project.members.all():
            if member.worker_id == worker.id:
                total_income += payments.get(member.id, Decimal("0.00"))
                break
        if project.status == "finished":
            finished_count += 1
        if project.amount is not None:
            project_amount_sum += Decimal(project.amount)
            amount_count += 1

    project_count = len(assigned_projects)
    performance_score_pct = (Decimal(finished_count) / Decimal(project_count) * Decimal("100")) if project_count else Decimal("0.00")
    performance_score_out_of_5 = (performance_score_pct / Decimal("100")) * Decimal("5")

    project_rows = []
    for item in assigned_projects:
        project_rows.append(
            {
                "project_name": item.title,
                "project_category": item.get_category_display(),
                "project_start_date": item.start_date.strftime("%Y-%m-%d"),
                "project_status": item.get_status_display(),
                "project_amount": item.amount or Decimal("0.00"),
            }
        )

    return {
        "department_name": dept.name,
        "worker_name": worker.name,
        "generated_at": timezone.localtime().strftime("%Y-%m-%d %H:%M:%S"),
        "worker_type": worker.get_worker_type_display(),
        "email": worker.email or "-",
        "department_role": worker.department_role,
        "posting": worker.posting,
        "date_of_join": worker.date_of_join.strftime("%Y-%m-%d"),
        "working_status": worker.get_working_status_display(),
        "project_count": project_count,
        "total_income": total_income,
        "performance_score_out_of_5": performance_score_out_of_5,
        "project_rows": project_rows,
    }


def generate_worker_csv_report(dept, worker):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        has_openpyxl = True
    except ImportError:
        has_openpyxl = False

    if not has_openpyxl:
        return JsonResponse(
            {"detail": "Excel generation dependency missing. Install openpyxl."},
            status=500,
        )

    report = build_worker_report_data(dept, worker)

    wb = Workbook()
    ws = wb.active
    ws.title = "Worker Report"

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    summary_rows = [
        ("Department Name", report["department_name"]),
        ("Worker Name", report["worker_name"]),
        ("Worker Type", report["worker_type"]),
        ("Email", report["email"]),
        ("Department Role", report["department_role"]),
        ("Posting", report["posting"]),
        ("Date Of Join", report["date_of_join"]),
        ("Working Status", report["working_status"]),
        ("Project Count", str(report["project_count"])),
        ("Total Income", f"₹{report['total_income']:,.2f}"),
        ("Performance Score (Out Of 5)", f"{report['performance_score_out_of_5']:.2f}/5"),
        ("Generated At", report["generated_at"]),
    ]

    row_idx = 1
    for label, value in summary_rows:
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)
        row_idx += 1

    row_idx += 1
    headers = ["Project Name", "Project Category", "Project Start Date", "Project Status", "Project Amount"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    row_idx += 1

    for item in report["project_rows"]:
        ws.cell(row=row_idx, column=1, value=item["project_name"])
        ws.cell(row=row_idx, column=2, value=item["project_category"])
        ws.cell(row=row_idx, column=3, value=item["project_start_date"])
        ws.cell(row=row_idx, column=4, value=item["project_status"])
        ws.cell(row=row_idx, column=5, value=f"₹{Decimal(item['project_amount']):,.2f}")
        for col_idx in range(1, 6):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="left", vertical="center")
        row_idx += 1

    ws.freeze_panes = "A14"

    filename = f"worker_{worker.id}_report.xlsx"
    output = BytesIO()
    wb.save(output)
    payload = output.getvalue()
    output.close()

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write(payload)
    return response


def generate_worker_pdf_report(dept, worker):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    except ImportError:
        return HttpResponse(
            "PDF generation dependency is missing. Install reportlab.",
            status=500,
            content_type="text/plain",
        )

    report = build_worker_report_data(dept, worker)
    filename = f"worker_{worker.id}_report.pdf"
    buffer = BytesIO()

    font_name = "Helvetica"
    bold_font_name = "Helvetica-Bold"
    for path in [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/calibri.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("ReportSans", path))
                font_name = "ReportSans"
                bold_font_name = "ReportSans"
                break
            except Exception:
                pass

    BRAND_DARK = colors.HexColor("#1A2B4A")
    BRAND_MID = colors.HexColor("#2E5FA3")
    BRAND_LIGHT = colors.HexColor("#E8F0FB")
    ACCENT = colors.HexColor("#F0A500")
    WHITE = colors.white
    GREY_TEXT = colors.HexColor("#4A4A4A")
    ROW_ALT = colors.HexColor("#EEF3FB")

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"{report['worker_name']} Worker Report",
        author=report["department_name"],
    )
    page_w = A4[0] - 20 * mm

    def style(name, **kwargs):
        return ParagraphStyle(name, **kwargs)

    getSampleStyleSheet()

    s_report_title = style(
        "ReportTitle",
        fontName=bold_font_name, fontSize=22,
        textColor=WHITE, alignment=TA_LEFT,
        spaceAfter=0, spaceBefore=0, leading=24,
    )
    s_report_subtitle = style(
        "ReportSubtitle",
        fontName=bold_font_name, fontSize=16,
        textColor=ACCENT, alignment=TA_RIGHT,
        spaceAfter=0, spaceBefore=0, leading=18,
    )
    s_section_heading = style(
        "SectionHeading",
        fontName=bold_font_name, fontSize=11,
        textColor=BRAND_DARK, spaceBefore=14, spaceAfter=4,
    )
    s_label = style(
        "Label",
        fontName=bold_font_name, fontSize=9,
        textColor=BRAND_DARK,
    )
    s_table_header = style(
        "TableHeader",
        fontName=bold_font_name, fontSize=9,
        textColor=WHITE, alignment=TA_LEFT,
    )
    s_table_cell = style(
        "TableCell",
        fontName=bold_font_name, fontSize=9,
        textColor=GREY_TEXT, leading=12,
    )
    s_table_cell_bold = style(
        "TableCellBold",
        fontName=bold_font_name, fontSize=8.5,
        textColor=BRAND_DARK, leading=12,
    )

    header_table = Table(
        [[
            Paragraph(f"{report['worker_name']} Report", s_report_title),
            Paragraph(report["department_name"], s_report_subtitle),
        ]],
        colWidths=[page_w * 0.7, page_w * 0.3],
    )
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), BRAND_DARK),
        ("TOPPADDING", (0, 0), (-1, -1), 14),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
        ("LEFTPADDING", (0, 0), (-1, -1), 16),
        ("RIGHTPADDING", (0, 0), (-1, -1), 16),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, 0), "LEFT"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))

    accent_bar = Table([[""]], colWidths=[page_w], rowHeights=[4])
    accent_bar.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), ACCENT),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    def summary_card(label, value):
        return [
            Paragraph(label, s_label),
            Paragraph(
                str(value),
                style(
                    "CardVal",
                    fontName=bold_font_name,
                    fontSize=12,
                    textColor=BRAND_MID,
                    spaceAfter=0,
                ),
            ),
        ]

    cards = [
        summary_card("Worker Type", report["worker_type"]),
        summary_card("Email", report["email"]),
        summary_card("Department Role", report["department_role"]),
        summary_card("Posting", report["posting"]),
        summary_card("Date Of Join", report["date_of_join"]),
        summary_card("Working Status", report["working_status"]),
        summary_card("Project Count", report["project_count"]),
        summary_card("Total Income", f"₹{report['total_income']:,.2f}"),
        summary_card("Performance Score (Out Of 5)", f"{report['performance_score_out_of_5']:.2f}/5"),
    ]
    summary_table = Table(
        [cards[0:3], cards[3:6], cards[6:9]],
        colWidths=[page_w / 3] * 3,
    )
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), BRAND_LIGHT),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LINEAFTER", (0, 0), (1, -1), 0.5, colors.HexColor("#B0C4E8")),
        ("LINEBELOW", (0, 0), (-1, 1), 0.5, colors.HexColor("#B0C4E8")),
    ]))

    rows = [[
        Paragraph("#", s_table_header),
        Paragraph("Project Name", s_table_header),
        Paragraph("Project Category", s_table_header),
        Paragraph("Project Start Date", s_table_header),
        Paragraph("Project Status", s_table_header),
        Paragraph("Project Amount", s_table_header),
    ]]

    for index, item in enumerate(report["project_rows"], start=1):
        rows.append([
            Paragraph(str(index), s_table_cell),
            Paragraph(item["project_name"], s_table_cell_bold),
            Paragraph(item["project_category"], s_table_cell),
            Paragraph(item["project_start_date"], s_table_cell),
            Paragraph(item["project_status"], s_table_cell),
            Paragraph(f"₹{Decimal(item['project_amount']):,.2f}", s_table_cell),
        ])

    row_bg_cmds = []
    for r in range(1, len(rows)):
        row_bg_cmds.append(("BACKGROUND", (0, r), (-1, r), ROW_ALT if r % 2 == 0 else WHITE))

    table = Table(
        rows,
        colWidths=[
            page_w * 0.05,
            page_w * 0.26,
            page_w * 0.17,
            page_w * 0.17,
            page_w * 0.15,
            page_w * 0.20,
        ],
        repeatRows=1,
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_MID),
        ("TOPPADDING", (0, 0), (-1, 0), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 1), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#D0DCF0")),
        ("LINEBELOW", (0, 0), (-1, 0), 0, WHITE),
        ("BOX", (0, 0), (-1, -1), 0.8, BRAND_MID),
        *row_bg_cmds,
    ]))

    footer_note = Paragraph(
        f"This report was automatically generated for "
        f"<b>{report['worker_name']}</b> in <b>{report['department_name']}</b> department.",
        style(
            "Footer",
            fontName=font_name,
            fontSize=8,
            textColor=colors.HexColor("#888888"),
            alignment=TA_CENTER,
        ),
    )

    def add_page_footer(canvas_obj, doc_obj):
        canvas_obj.saveState()
        canvas_obj.setFont(font_name, 8)
        canvas_obj.setFillColor(colors.HexColor("#AAAAAA"))
        canvas_obj.drawString(
            10 * mm, 10 * mm,
            f"Worker Report  |  {report['worker_name']}"
        )
        canvas_obj.drawRightString(
            A4[0] - 10 * mm, 10 * mm,
            f"Page {doc_obj.page}"
        )
        canvas_obj.setStrokeColor(colors.HexColor("#DDDDDD"))
        canvas_obj.line(10 * mm, 13 * mm, A4[0] - 10 * mm, 13 * mm)
        canvas_obj.restoreState()

    story = [
        header_table,
        accent_bar,
        Spacer(1, 10),
        Paragraph("Worker Summary", s_section_heading),
        summary_table,
        Spacer(1, 14),
        HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#D0DCF0")),
        Spacer(1, 6),
        Paragraph("Working Project Details", s_section_heading),
        table,
        Spacer(1, 20),
        HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#DDDDDD")),
        Spacer(1, 6),
        footer_note,
    ]

    doc.build(story, onFirstPage=add_page_footer, onLaterPages=add_page_footer)

    pdf_bytes = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write(pdf_bytes)
    return response
