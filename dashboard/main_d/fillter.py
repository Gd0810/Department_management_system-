from decimal import Decimal
from io import BytesIO
import os

from django.http import HttpResponse, JsonResponse
from django.db.models import Sum
from django.utils import timezone


def _calculate_project_payments(project):
    if not project.amount:
        return {}

    members = list(project.members.select_related("worker"))
    gold = [m for m in members if m.contribution == "gold"]
    silver = [m for m in members if m.contribution == "silver"]
    copper = [m for m in members if m.contribution == "copper"]
    total_amount = Decimal(project.amount)
    payments = {}

    if gold and not silver and not copper:
        share = total_amount / len(gold)
        for member in gold:
            payments[member.id] = share
    elif gold and silver and not copper:
        gold_total = total_amount * Decimal("0.60")
        silver_total = total_amount * Decimal("0.40")
        for member in gold:
            payments[member.id] = gold_total / len(gold)
        for member in silver:
            payments[member.id] = silver_total / len(silver)
    elif gold and copper and not silver:
        gold_total = total_amount * Decimal("0.70")
        copper_total = total_amount * Decimal("0.30")
        for member in gold:
            payments[member.id] = gold_total / len(gold)
        for member in copper:
            payments[member.id] = copper_total / len(copper)
    else:
        weight_map = {"gold": 3, "silver": 2, "copper": 1}
        total_weight = sum(weight_map[m.contribution] for m in members)
        for member in members:
            share = (Decimal(weight_map[member.contribution]) / Decimal(total_weight)) * total_amount
            payments[member.id] = share

    return payments


def build_main_filter_report_data(dept, start_date, end_date, range_key):
    projects = (
        dept.projects.filter(start_date__gte=start_date, start_date__lte=end_date)
        .prefetch_related("members__worker")
        .order_by("-start_date", "-id")
    )
    workers = dept.workers.all().order_by("name")

    filtered_income = projects.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
    filtered_project_count = projects.count()

    project_rows = []
    for project in projects:
        assigned_worker_names = ", ".join(sorted({m.worker.name for m in project.members.all()})) or "-"
        project_rows.append(
            {
                "project_name": project.title,
                "start_date": project.start_date.strftime("%Y-%m-%d"),
                "category": project.get_category_display(),
                "status": project.get_status_display(),
                "amount": Decimal(project.amount or Decimal("0.00")),
                "assigned_workers": assigned_worker_names,
            }
        )

    worker_income_map = {worker.id: Decimal("0.00") for worker in workers}
    for project in projects:
        payments = _calculate_project_payments(project)
        for member in project.members.all():
            worker_income_map[member.worker_id] = (
                worker_income_map.get(member.worker_id, Decimal("0.00")) + payments.get(member.id, Decimal("0.00"))
            )

    worker_rows = []
    for worker in workers:
        income = worker_income_map.get(worker.id, Decimal("0.00"))
        if income <= 0:
            continue
        worker_rows.append(
            {
                "name": worker.name,
                "email": worker.email or "-",
                "date_of_join": worker.date_of_join.strftime("%Y-%m-%d"),
                "posting": worker.posting,
                "income_by_user": income,
            }
        )

    income_label = "Month Income" if range_key == "month" else "Filtered Income"
    count_label = "Project Count" if range_key == "month" else "Filtered Project Count"

    return {
        "department_name": dept.name,
        "range_key": range_key,
        "start_date": start_date.strftime("%Y-%m-%d"),
        "end_date": end_date.strftime("%Y-%m-%d"),
        "income_label": income_label,
        "count_label": count_label,
        "filtered_income": filtered_income,
        "filtered_project_count": filtered_project_count,
        "project_rows": project_rows,
        "worker_rows": worker_rows,
        "generated_at": timezone.localtime().strftime("%Y-%m-%d %H:%M:%S"),
    }


def generate_main_filter_csv_report(dept, start_date, end_date, range_key):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        has_openpyxl = True
    except ImportError:
        has_openpyxl = False

    if not has_openpyxl:
        return JsonResponse(
            {"detail": "Excel generation dependency missing. Install openpyxl."},
            status=500,
        )

    report = build_main_filter_report_data(dept, start_date, end_date, range_key)
    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered Report"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 40

    summary_rows = [
        ("Department Name", report["department_name"]),
        ("Date Range", f"{report['start_date']} to {report['end_date']}"),
        (report["income_label"], f"Rs {report['filtered_income']:,.2f}"),
        (report["count_label"], str(report["filtered_project_count"])),
        ("Generated At", report["generated_at"]),
    ]

    row_idx = 1
    for label, value in summary_rows:
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)
        row_idx += 1

    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Table 1: Filtered Project List").font = Font(bold=True)
    row_idx += 1
    headers_1 = ["#", "Project Name", "Start Date", "Category", "Status", "Amount", "Assigned Workers"]
    for col_idx, header in enumerate(headers_1, start=1):
        ws.cell(row=row_idx, column=col_idx, value=header).font = Font(bold=True)
    row_idx += 1

    for index, item in enumerate(report["project_rows"], start=1):
        ws.cell(row=row_idx, column=1, value=index)
        ws.cell(row=row_idx, column=2, value=item["project_name"])
        ws.cell(row=row_idx, column=3, value=item["start_date"])
        ws.cell(row=row_idx, column=4, value=item["category"])
        ws.cell(row=row_idx, column=5, value=item["status"])
        ws.cell(row=row_idx, column=6, value=f"Rs {Decimal(item['amount']):,.2f}")
        ws.cell(row=row_idx, column=7, value=item["assigned_workers"])
        for col_idx in range(1, 8):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="left", vertical="center", wrap_text=(col_idx in [2, 7]))
        row_idx += 1

    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Table 2: Filtered Worker Contribution List").font = Font(bold=True)
    row_idx += 1
    headers_2 = ["#", "Name", "Email", "Date Of Join", "Posting", "Income By User"]
    for col_idx, header in enumerate(headers_2, start=1):
        ws.cell(row=row_idx, column=col_idx, value=header).font = Font(bold=True)
    row_idx += 1

    for index, item in enumerate(report["worker_rows"], start=1):
        ws.cell(row=row_idx, column=1, value=index)
        ws.cell(row=row_idx, column=2, value=item["name"])
        ws.cell(row=row_idx, column=3, value=item["email"])
        ws.cell(row=row_idx, column=4, value=item["date_of_join"])
        ws.cell(row=row_idx, column=5, value=item["posting"])
        ws.cell(row=row_idx, column=6, value=f"Rs {Decimal(item['income_by_user']):,.2f}")
        for col_idx in range(1, 7):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="left", vertical="center", wrap_text=(col_idx in [2, 3, 5]))
        row_idx += 1

    output = BytesIO()
    wb.save(output)
    payload = output.getvalue()
    output.close()

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="main_filtered_report.xlsx"'
    response.write(payload)
    return response


def generate_main_filter_pdf_report(dept, start_date, end_date, range_key):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    except ImportError:
        return HttpResponse(
            "PDF generation dependency is missing. Install reportlab.",
            status=500,
            content_type="text/plain",
        )

    report = build_main_filter_report_data(dept, start_date, end_date, range_key)

    font_name = "Helvetica"
    bold_font_name = "Helvetica-Bold"
    for path in [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/calibri.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("ReportSans", path))
                font_name = "ReportSans"
                bold_font_name = "ReportSans"
                break
            except Exception:
                pass

    BRAND_DARK = colors.HexColor("#1A2B4A")
    BRAND_MID = colors.HexColor("#2E5FA3")
    BRAND_LIGHT = colors.HexColor("#E8F0FB")
    ACCENT = colors.HexColor("#F0A500")
    WHITE = colors.white
    GREY_TEXT = colors.HexColor("#4A4A4A")
    ROW_ALT = colors.HexColor("#EEF3FB")

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Filtered Overall Report",
        author=report["department_name"],
    )
    page_w = A4[0] - 20 * mm

    def style(name, **kwargs):
        return ParagraphStyle(name, **kwargs)

    getSampleStyleSheet()

    s_report_title = style("ReportTitle", fontName=bold_font_name, fontSize=22, textColor=WHITE, alignment=TA_LEFT, leading=24)
    s_report_subtitle = style("ReportSubtitle", fontName=bold_font_name, fontSize=14, textColor=ACCENT, alignment=TA_RIGHT, leading=18)
    s_section_heading = style("SectionHeading", fontName=bold_font_name, fontSize=11, textColor=BRAND_DARK, spaceBefore=14, spaceAfter=4)
    s_label = style("Label", fontName=bold_font_name, fontSize=9, textColor=BRAND_DARK)
    s_table_header = style("TableHeader", fontName=bold_font_name, fontSize=8.8, textColor=WHITE, alignment=TA_LEFT)
    s_table_cell = style("TableCell", fontName=bold_font_name, fontSize=8.5, textColor=GREY_TEXT, leading=11)
    s_table_cell_bold = style("TableCellBold", fontName=bold_font_name, fontSize=8.5, textColor=BRAND_DARK, leading=11)

    header_table = Table(
        [[
            Paragraph(f"{report['department_name']} report", s_report_title),
            Paragraph(f"{report['start_date']} to {report['end_date']}", s_report_subtitle),
        ]],
        colWidths=[page_w * 0.7, page_w * 0.3],
    )
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), BRAND_DARK),
        ("TOPPADDING", (0, 0), (-1, -1), 14),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
        ("LEFTPADDING", (0, 0), (-1, -1), 16),
        ("RIGHTPADDING", (0, 0), (-1, -1), 16),
    ]))

    accent_bar = Table([[""]], colWidths=[page_w], rowHeights=[4])
    accent_bar.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), ACCENT)]))

    def summary_card(label, value):
        return [Paragraph(label, s_label), Paragraph(str(value), style("CardVal", fontName=bold_font_name, fontSize=12, textColor=BRAND_MID))]

    summary_table = Table(
        [[
            summary_card(report["income_label"], f"Rs {report['filtered_income']:,.2f}"),
            summary_card(report["count_label"], report["filtered_project_count"]),
            summary_card("Generated At", report["generated_at"]),
        ]],
        colWidths=[page_w / 3] * 3,
    )
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), BRAND_LIGHT),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("LINEAFTER", (0, 0), (1, 0), 0.5, colors.HexColor("#B0C4E8")),
    ]))

    project_rows = [[
        Paragraph("#", s_table_header),
        Paragraph("Project Name", s_table_header),
        Paragraph("Start Date", s_table_header),
        Paragraph("Category", s_table_header),
        Paragraph("Status", s_table_header),
        Paragraph("Amount", s_table_header),
        Paragraph("Assigned Workers", s_table_header),
    ]]
    for idx, item in enumerate(report["project_rows"], start=1):
        project_rows.append([
            Paragraph(str(idx), s_table_cell),
            Paragraph(item["project_name"], s_table_cell_bold),
            Paragraph(item["start_date"], s_table_cell),
            Paragraph(item["category"], s_table_cell),
            Paragraph(item["status"], s_table_cell),
            Paragraph(f"Rs {Decimal(item['amount']):,.2f}", s_table_cell),
            Paragraph(item["assigned_workers"], s_table_cell),
        ])

    project_bg_cmds = [("BACKGROUND", (0, r), (-1, r), ROW_ALT if r % 2 == 0 else WHITE) for r in range(1, len(project_rows))]
    project_table = Table(
        project_rows,
        colWidths=[page_w * 0.05, page_w * 0.20, page_w * 0.11, page_w * 0.11, page_w * 0.11, page_w * 0.11, page_w * 0.31],
        repeatRows=1,
    )
    project_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_MID),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("BOX", (0, 0), (-1, -1), 0.8, BRAND_MID),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#D0DCF0")),
        *project_bg_cmds,
    ]))

    worker_rows = [[
        Paragraph("#", s_table_header),
        Paragraph("Name", s_table_header),
        Paragraph("Email", s_table_header),
        Paragraph("Date Of Join", s_table_header),
        Paragraph("Posting", s_table_header),
        Paragraph("Income By User", s_table_header),
    ]]
    for idx, item in enumerate(report["worker_rows"], start=1):
        worker_rows.append([
            Paragraph(str(idx), s_table_cell),
            Paragraph(item["name"], s_table_cell_bold),
            Paragraph(item["email"], s_table_cell),
            Paragraph(item["date_of_join"], s_table_cell),
            Paragraph(item["posting"], s_table_cell),
            Paragraph(f"Rs {Decimal(item['income_by_user']):,.2f}", s_table_cell),
        ])

    worker_bg_cmds = [("BACKGROUND", (0, r), (-1, r), ROW_ALT if r % 2 == 0 else WHITE) for r in range(1, len(worker_rows))]
    worker_table = Table(
        worker_rows,
        colWidths=[page_w * 0.06, page_w * 0.17, page_w * 0.30, page_w * 0.14, page_w * 0.16, page_w * 0.17],
        repeatRows=1,
    )
    worker_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_MID),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("BOX", (0, 0), (-1, -1), 0.8, BRAND_MID),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#D0DCF0")),
        *worker_bg_cmds,
    ]))

    footer_note = Paragraph(
        f"Filtered report generated for <b>{report['department_name']}</b> ({report['start_date']} to {report['end_date']}).",
        style("Footer", fontName=font_name, fontSize=8, textColor=colors.HexColor("#888888"), alignment=TA_CENTER),
    )

    def add_page_footer(canvas_obj, doc_obj):
        canvas_obj.saveState()
        canvas_obj.setFont(font_name, 8)
        canvas_obj.setFillColor(colors.HexColor("#AAAAAA"))
        canvas_obj.drawString(10 * mm, 10 * mm, f"Filtered Overall Report  |  {report['department_name']}")
        canvas_obj.drawRightString(A4[0] - 10 * mm, 10 * mm, f"Page {doc_obj.page}")
        canvas_obj.restoreState()

    story = [
        header_table,
        accent_bar,
        Spacer(1, 10),
        Paragraph("Filtered Summary", s_section_heading),
        summary_table,
        Spacer(1, 12),
        HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#D0DCF0")),
        Spacer(1, 6),
        Paragraph("Project List (Filtered)", s_section_heading),
        project_table,
        Spacer(1, 12),
        Paragraph("Worker Contribution List (Filtered)", s_section_heading),
        worker_table,
        Spacer(1, 12),
        footer_note,
    ]
    doc.build(story, onFirstPage=add_page_footer, onLaterPages=add_page_footer)

    payload = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="main_filtered_report.pdf"'
    response.write(payload)
    return response
