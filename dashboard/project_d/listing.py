from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse, JsonResponse

from ..models import Project


def _apply_listing_filters(queryset, query_params):
    q = (query_params.get("q") or "").strip()
    month_value = (query_params.get("month") or "").strip()
    year_value = (query_params.get("year") or "").strip()
    status_value = (query_params.get("status") or "").strip()

    if q:
        queryset = queryset.filter(title__icontains=q)

    if month_value:
        try:
            month_year, month_num = month_value.split("-")
            queryset = queryset.filter(
                start_date__year=int(month_year),
                start_date__month=int(month_num),
            )
        except (TypeError, ValueError):
            pass

    if year_value:
        try:
            queryset = queryset.filter(start_date__year=int(year_value))
        except (TypeError, ValueError):
            pass

    valid_statuses = {choice[0] for choice in Project.PROJECT_STATUS}
    if status_value in valid_statuses:
        queryset = queryset.filter(status=status_value)

    return queryset


def _listing_rows(projects):
    rows = []
    for project in projects:
        worker_names = ", ".join(
            project.members.all().values_list("worker__name", flat=True)
        ) or "No workers assigned"
        rows.append(
            {
                "project_name": project.title,
                "project_category": project.get_category_display(),
                "start_date": project.start_date.strftime("%Y-%m-%d"),
                "status": project.get_status_display(),
                "amount": project.amount or Decimal("0.00"),
                "assigned_workers": worker_names,
            }
        )
    return rows


def _filtered_projects_for_report(dept, category_key, query_params):
    queryset = (
        dept.projects.filter(category=category_key)
        .prefetch_related("members__worker")
        .order_by("-start_date", "-id")
    )
    return _apply_listing_filters(queryset, query_params)


def generate_project_listing_excel_report(dept, category_key, query_params):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        has_openpyxl = True
    except ImportError:
        has_openpyxl = False

    projects = _filtered_projects_for_report(dept, category_key, query_params)
    rows = _listing_rows(projects)

    if not has_openpyxl:
        return JsonResponse(
            {"detail": "Excel generation dependency missing. Install openpyxl."},
            status=500,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Project Listing"

    headers = [
        "Project Name",
        "Project Category",
        "Start Date",
        "Status",
        "Amount",
        "Assigned Workers",
    ]
    widths = [34, 18, 14, 14, 12, 48]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    row_idx = 2
    for item in rows:
        ws.cell(row=row_idx, column=1, value=item["project_name"])
        ws.cell(row=row_idx, column=2, value=item["project_category"])
        ws.cell(row=row_idx, column=3, value=item["start_date"])
        ws.cell(row=row_idx, column=4, value=item["status"])
        ws.cell(row=row_idx, column=5, value=f"\u20B9{Decimal(item['amount']):,.2f}")
        ws.cell(row=row_idx, column=6, value=item["assigned_workers"])
        for col_idx in range(1, 7):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=(col_idx == 6),
            )
        row_idx += 1

    ws.freeze_panes = "A2"

    filename = f"{category_key}_project_listing.xlsx"
    output = BytesIO()
    wb.save(output)
    payload = output.getvalue()
    output.close()

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write(payload)
    return response


def generate_project_listing_pdf_report(dept, category_key, query_params):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
        from reportlab.lib.enums import TA_LEFT
    except ImportError:
        return HttpResponse(
            "PDF generation dependency is missing. Install reportlab.",
            status=500,
            content_type="text/plain",
        )

    projects = _filtered_projects_for_report(dept, category_key, query_params)
    rows = _listing_rows(projects)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"{category_key.title()} Project Listing",
    )

    table_header_style = ParagraphStyle(
        "th",
        fontName="Helvetica-Bold",
        fontSize=9,
        textColor=colors.white,
        alignment=TA_LEFT,
    )
    table_cell_style = ParagraphStyle(
        "td",
        fontName="Helvetica",
        fontSize=8.5,
        textColor=colors.HexColor("#1f2937"),
        alignment=TA_LEFT,
    )

    data = [[
        Paragraph("Project Name", table_header_style),
        Paragraph("Project Category", table_header_style),
        Paragraph("Start Date", table_header_style),
        Paragraph("Status", table_header_style),
        Paragraph("Amount", table_header_style),
        Paragraph("Assigned Workers", table_header_style),
    ]]
    for item in rows:
        data.append([
            Paragraph(item["project_name"], table_cell_style),
            Paragraph(item["project_category"], table_cell_style),
            Paragraph(item["start_date"], table_cell_style),
            Paragraph(item["status"], table_cell_style),
            Paragraph(f"\u20B9{Decimal(item['amount']):,.2f}", table_cell_style),
            Paragraph(item["assigned_workers"], table_cell_style),
        ])

    usable_width = A4[0] - (20 * mm)
    col_widths = [
        usable_width * 0.23,
        usable_width * 0.17,
        usable_width * 0.12,
        usable_width * 0.12,
        usable_width * 0.10,
        usable_width * 0.26,
    ]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5FA3")),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d1d5db")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))

    doc.build([table])
    payload = buffer.getvalue()
    buffer.close()

    filename = f"{category_key}_project_listing.pdf"
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write(payload)
    return response
