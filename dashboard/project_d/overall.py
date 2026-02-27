import csv
import os
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from django.db.models import Sum
from django.utils import timezone

from ..models import Project


def _category_label(category_key):
    return dict(Project.PROJECT_CATEGORY).get(category_key, category_key.title())


def build_category_report_data(dept, category_key):
    category_label = _category_label(category_key)
    projects = (
        dept.projects.filter(category=category_key)
        .prefetch_related("members__worker")
        .order_by("-start_date", "-id")
    )
    overall_income = (
        projects.aggregate(total_income=Sum("amount"))["total_income"]
        or Decimal("0.00")
    )
    overall_project_count = projects.count()

    rows = []
    for project in projects:
        worker_names = ", ".join(
            project.members.all()
            .values_list("worker__name", flat=True)
        ) or "No workers assigned"
        rows.append(
            {
                "project_name": project.title,
                "start_date": project.start_date.strftime("%Y-%m-%d"),
                "status": project.get_status_display(),
                "amount": project.amount or Decimal("0.00"),
                "worker_names": worker_names,
            }
        )

    return {
        "department_name": dept.name,
        "category_key": category_key,
        "category_label": category_label,
        "overall_income": overall_income,
        "overall_project_count": overall_project_count,
        "projects": rows,
        "generated_at": timezone.localtime().strftime("%Y-%m-%d %H:%M:%S"),
    }



def generate_category_csv_report(dept, category_key):
    """Generate minimal Excel report with proper table alignment and no UI styling."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        has_openpyxl = True
    except ImportError:
        has_openpyxl = False

    report = build_category_report_data(dept, category_key)

    if not has_openpyxl:
        filename = f"{category_key}_projects_report.csv"
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        writer = csv.writer(response, lineterminator="\n")
        writer.writerow(["Department Name", report["department_name"]])
        writer.writerow(["Category", report["category_label"]])
        writer.writerow(["Total Income", f"\u20B9{report['overall_income']:,.2f}"])
        writer.writerow(["Total Projects", report["overall_project_count"]])
        writer.writerow(["Generated At", report["generated_at"]])
        writer.writerow([])
        writer.writerow(["Project Name", "Start Date", "Status", "Amount", "Assigned Workers"])
        for row in report["projects"]:
            writer.writerow([
                row["project_name"],
                row["start_date"],
                row["status"],
                f"\u20B9{Decimal(row['amount']):,.2f}",
                row["worker_names"],
            ])
        return response

    wb = Workbook()
    ws = wb.active
    ws.title = "Projects Report"
    ws.sheet_view.showGridLines = True

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 48

    row_idx = 1
    summary_rows = [
        ("Department Name", report["department_name"]),
        ("Category", report["category_label"]),
        ("Total Income", f"\u20B9{report['overall_income']:,.2f}"),
        ("Total Projects", str(report["overall_project_count"])),
        ("Generated At", report["generated_at"]),
    ]

    for label, value in summary_rows:
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)
        row_idx += 1

    row_idx += 1
    headers = ["Project Name", "Start Date", "Status", "Amount", "Assigned Workers"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    table_first_data_row = row_idx + 1
    row_idx += 1

    for item in report["projects"]:
        ws.cell(row=row_idx, column=1, value=item["project_name"])
        ws.cell(row=row_idx, column=2, value=item["start_date"])
        ws.cell(row=row_idx, column=3, value=item["status"])
        ws.cell(row=row_idx, column=4, value=f"\u20B9{Decimal(item['amount']):,.2f}")
        ws.cell(row=row_idx, column=5, value=item["worker_names"])
        for col_idx in range(1, 6):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=(col_idx == 5),
            )
        row_idx += 1

    ws.freeze_panes = f"A{table_first_data_row}"

    filename = f"{category_key}_projects_report.xlsx"
    output = BytesIO()
    wb.save(output)
    payload = output.getvalue()
    output.close()

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write(payload)
    return response

def generate_category_pdf_report(dept, category_key):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm, cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable, KeepTogether
        )
        from reportlab.platypus.flowables import HRFlowable
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    except ImportError:
        return HttpResponse(
            "PDF generation dependency is missing. Install reportlab.",
            status=500,
            content_type="text/plain",
        )

    report = build_category_report_data(dept, category_key)
    filename = f"{category_key}_projects_report.pdf"
    buffer = BytesIO()

    # Use a Unicode TTF font so ₹ renders correctly in PDF.
    font_name = "Helvetica"
    bold_font_name = "Helvetica-Bold"
    for path in [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/calibri.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("ReportSans", path))
                font_name = "ReportSans"
                bold_font_name = "ReportSans"
                break
            except Exception:
                pass

    # ── Colour palette ──────────────────────────────────────────────────────
    BRAND_DARK   = colors.HexColor("#1A2B4A")   # deep navy  – header bg
    BRAND_MID    = colors.HexColor("#2E5FA3")   # medium blue – accent
    BRAND_LIGHT  = colors.HexColor("#E8F0FB")   # pale blue  – alt row / summary
    ACCENT       = colors.HexColor("#F0A500")   # amber      – highlight bar
    WHITE        = colors.white
    GREY_TEXT    = colors.HexColor("#4A4A4A")
    LIGHT_GREY   = colors.HexColor("#F5F5F5")
    ROW_ALT      = colors.HexColor("#EEF3FB")

    # ── Document setup ──────────────────────────────────────────────────────
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"{report['category_label']} Projects Report",
        author=report.get("department_name", ""),
    )
    page_w = A4[0] - 20 * mm   # usable width

    # ── Styles ───────────────────────────────────────────────────────────────
    base = getSampleStyleSheet()

    def style(name, **kw):
        return ParagraphStyle(name, **kw)

    S_REPORT_TITLE = style(
        "ReportTitle",
        fontName=bold_font_name, fontSize=22,
        textColor=WHITE, alignment=TA_LEFT,
        spaceAfter=0, spaceBefore=0, leading=24,
    )
    S_REPORT_SUBTITLE = style(
        "ReportSubtitle",
        fontName=bold_font_name, fontSize=16,
        textColor=colors.HexColor("#F0A500"), alignment=TA_RIGHT,
        spaceAfter=0, spaceBefore=0, leading=18,
    )
    S_SECTION_HEADING = style(
        "SectionHeading",
        fontName=bold_font_name, fontSize=11,
        textColor=BRAND_DARK, spaceBefore=14, spaceAfter=4,
    )
    S_BODY = style(
        "Body",
        fontName=font_name, fontSize=9,
        textColor=GREY_TEXT, leading=14,
    )
    S_LABEL = style(
        "Label",
        fontName=bold_font_name, fontSize=9,
        textColor=BRAND_DARK,
    )
    S_TABLE_HEADER = style(
        "TableHeader",
        fontName=bold_font_name, fontSize=9,
        textColor=WHITE, alignment=TA_LEFT,
    )
    S_TABLE_CELL = style(
        "TableCell",
        fontName=bold_font_name, fontSize=9,
        textColor=GREY_TEXT, leading=12,
    )
    S_TABLE_CELL_BOLD = style(
        "TableCellBold",
        fontName=bold_font_name, fontSize=8.5,
        textColor=BRAND_DARK, leading=12,
    )

    # ── Header banner ────────────────────────────────────────────────────────
    header_table = Table(
        [[
            Paragraph(report["category_label"] + " Projects Report", S_REPORT_TITLE),
            Paragraph(report["department_name"], S_REPORT_SUBTITLE),
        ]],
        colWidths=[page_w * 0.7, page_w * 0.3],
    )
    header_table.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), colors.HexColor("#1A2B4A")),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [BRAND_DARK]),
        ("TOPPADDING",   (0, 0), (-1, -1), 14),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 14),
        ("LEFTPADDING",  (0, 0), (-1, -1), 16),
        ("RIGHTPADDING", (0, 0), (-1, -1), 16),
        ("ROUNDEDCORNERS", [6]),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",        (0, 0), (0, 0),   "LEFT"),
        ("ALIGN",        (1, 0), (1, 0),   "RIGHT"),
    ]))

    # Amber accent bar under the header
    accent_bar = Table([[""]], colWidths=[page_w], rowHeights=[4])
    accent_bar.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), ACCENT),
        ("TOPPADDING",   (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 0),
    ]))

    # ── Summary cards ────────────────────────────────────────────────────────
    def summary_card(label, value):
        return [
            Paragraph(label, S_LABEL),
            Paragraph(str(value), style("CardVal", fontName=bold_font_name,
                                        fontSize=16, textColor=BRAND_MID,
                                        spaceAfter=0)),
        ]

    card_data = [
        summary_card("Total Income",   f"₹{report['overall_income']:,.2f}"),
        summary_card("Total Projects", str(report["overall_project_count"])),
        summary_card("Generated At",   report["generated_at"]),
    ]
    summary_table = Table(
        [card_data],
        colWidths=[page_w / 3] * 3,
    )
    summary_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), BRAND_LIGHT),
        ("TOPPADDING",    (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ("LEFTPADDING",   (0, 0), (-1, -1), 14),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 14),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("LINEAFTER",     (0, 0), (1, 0),   0.5, colors.HexColor("#B0C4E8")),
        ("ROUNDEDCORNERS", [4]),
    ]))

    # ── Projects table ────────────────────────────────────────────────────────
    col_headers = ["#", "Project Name", "Start Date", "Status", "Amount", "Assigned Workers"]
    col_widths  = [
        page_w * 0.05,
        page_w * 0.28,
        page_w * 0.13,
        page_w * 0.12,
        page_w * 0.12,
        page_w * 0.30,
    ]

    STATUS_COLORS = {
        "active":     colors.HexColor("#1E8C45"),
        "completed":  BRAND_MID,
        "on hold":    colors.HexColor("#C0392B"),
        "pending":    colors.HexColor("#D68910"),
    }

    def status_pill(text):
        colour = STATUS_COLORS.get(text.strip().lower(), GREY_TEXT)
        return Paragraph(
            f'<font color="{colour.hexval() if hasattr(colour,"hexval") else colour}">'
            f'<b>{text[:18]}</b></font>',
            S_TABLE_CELL,
        )

    rows = [[
        Paragraph(h, S_TABLE_HEADER) for h in col_headers
    ]]
    for i, item in enumerate(report["projects"], start=1):
        rows.append([
            Paragraph(str(i), S_TABLE_CELL),
            Paragraph(item["project_name"][:50], S_TABLE_CELL_BOLD),
            Paragraph(item["start_date"], S_TABLE_CELL),
            status_pill(item["status"]),
            Paragraph(f"₹{Decimal(item['amount']):,.2f}", S_TABLE_CELL),
            Paragraph(item["worker_names"][:120], S_TABLE_CELL),
        ])

    # Alternating row colours
    row_count = len(rows)
    row_bg_cmds = []
    for r in range(1, row_count):
        bg = ROW_ALT if r % 2 == 0 else WHITE
        row_bg_cmds.append(("BACKGROUND", (0, r), (-1, r), bg))

    projects_table = Table(rows, colWidths=col_widths, repeatRows=1)
    projects_table.setStyle(TableStyle([
        # Header row
        ("BACKGROUND",    (0, 0), (-1, 0),  BRAND_MID),
        ("TOPPADDING",    (0, 0), (-1, 0),  9),
        ("BOTTOMPADDING", (0, 0), (-1, 0),  9),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        # Data rows
        ("TOPPADDING",    (0, 1), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 7),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        # Grid
        ("LINEBELOW",     (0, 0), (-1, -1), 0.3, colors.HexColor("#D0DCF0")),
        ("LINEBELOW",     (0, 0), (-1,  0), 0,   WHITE),
        # Outer border
        ("BOX",           (0, 0), (-1, -1), 0.8, BRAND_MID),
        *row_bg_cmds,
    ]))

    # ── Footer note ──────────────────────────────────────────────────────────
    footer_note = Paragraph(
        f"This report was automatically generated for the "
        f"<b>{report['department_name']}</b> department. "
        f"All figures are indicative and subject to change.",
        style("Footer", fontName=font_name, fontSize=8,
              textColor=colors.HexColor("#888888"),
              alignment=TA_CENTER),
    )

    # ── Page footer callback ─────────────────────────────────────────────────
    def add_page_footer(canvas_obj, doc_obj):
        canvas_obj.saveState()
        canvas_obj.setFont(font_name, 8)
        canvas_obj.setFillColor(colors.HexColor("#AAAAAA"))
        canvas_obj.drawString(
            10 * mm, 10 * mm,
            f"{report['category_label']} Projects Report  |  {report['department_name']}"
        )
        canvas_obj.drawRightString(
            A4[0] - 10 * mm, 10 * mm,
            f"Page {doc_obj.page}"
        )
        canvas_obj.setStrokeColor(colors.HexColor("#DDDDDD"))
        canvas_obj.line(10 * mm, 13 * mm, A4[0] - 10 * mm, 13 * mm)
        canvas_obj.restoreState()

    # ── Assemble story ───────────────────────────────────────────────────────
    story = [
        header_table,
        accent_bar,
        Spacer(1, 10),
        Paragraph("Summary Overview", S_SECTION_HEADING),
        summary_table,
        Spacer(1, 14),
        HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#D0DCF0")),
        Spacer(1, 6),
        Paragraph("Project Details", S_SECTION_HEADING),
        projects_table,
        Spacer(1, 20),
        HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#DDDDDD")),
        Spacer(1, 6),
        footer_note,
    ]

    doc.build(story, onFirstPage=add_page_footer, onLaterPages=add_page_footer)

    pdf_bytes = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write(pdf_bytes)
    return response


